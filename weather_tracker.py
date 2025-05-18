import requests
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# Параметри
city_name = "Skopje"
latitude = 41.9981
longitude = 21.4254
start_date = "2023-01-01"
end_date = datetime.today().strftime('%Y-%m-%d')

# URL за историски податоци
url_history = (
    f"https://archive-api.open-meteo.com/v1/archive?"
    f"latitude={latitude}&longitude={longitude}&start_date={start_date}&end_date={end_date}"
    f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum"
    f"&timezone=Europe%2FBelgrade"
)

response_history = requests.get(url_history)
data_history = response_history.json()
df_history = pd.DataFrame(data_history["daily"])

# Преименувај ја колоната time во datum и остави само датум без време
df_history.rename(columns={"time": "datum"}, inplace=True)
df_history["datum"] = pd.to_datetime(df_history["datum"]).dt.date

# Подреди од најнов кон најстар
df_history.sort_values("datum", ascending=False, inplace=True)
df_history.reset_index(drop=True, inplace=True)

# URL за прогноза
url_forecast = (
    f"https://api.open-meteo.com/v1/forecast?"
    f"latitude={latitude}&longitude={longitude}"
    f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum"
    f"&timezone=Europe%2FBelgrade&forecast_days=16"
)

response_forecast = requests.get(url_forecast)
data_forecast = response_forecast.json()

if "daily" not in data_forecast:
    print("ГРЕШКА: Не може да се добијат прогностички податоци.")
    df_forecast = pd.DataFrame()
else:
    df_forecast = pd.DataFrame(data_forecast["daily"])
    df_forecast.rename(columns={"time": "datum"}, inplace=True)
    df_forecast["datum"] = pd.to_datetime(df_forecast["datum"]).dt.date
    df_forecast.sort_values("datum", inplace=True)

# Име на Excel фајл
filename = f"Weather_{city_name}_{start_date}_to_{end_date}.xlsx"

# Запиши во Excel
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df_history.to_excel(writer, sheet_name="2023-2025", index=False)
    if not df_forecast.empty:
        df_forecast.to_excel(writer, sheet_name="Forecast", index=False)

    # Sheet за споредба на исти датуми по години (температура)
    df_history_sorted = df_history.copy()
    df_history_sorted["year"] = pd.to_datetime(df_history_sorted["datum"]).dt.year
    df_history_sorted["month_day"] = pd.to_datetime(df_history_sorted["datum"]).dt.strftime("%m-%d")

    pivot_temp_max = df_history_sorted.pivot(index="month_day", columns="year", values="temperature_2m_max")
    pivot_temp_min = df_history_sorted.pivot(index="month_day", columns="year", values="temperature_2m_min")
    pivot_precip = df_history_sorted.pivot(index="month_day", columns="year", values="precipitation_sum")

    pivot_temp_max.to_excel(writer, sheet_name="Compare Tmax")
    pivot_temp_min.to_excel(writer, sheet_name="Compare Tmin")
    pivot_precip.to_excel(writer, sheet_name="Compare Rain")

# Бојање по години во sheet "2023-2025"
wb = load_workbook(filename)
ws = wb["2023-2025"]

color_map = {
    2023: "FFC7CE",  # светло црвена
    2024: "C6EFCE",  # светло зелена
    2025: "FFEB9C",  # светло жолта
}

date_col = 1  # прва колона "datum"

for row in ws.iter_rows(min_row=2):
    cell = row[date_col - 1]
    if cell.value:
        try:
            year = datetime.strptime(str(cell.value), "%Y-%m-%d").year
        except:
            continue
        if year in color_map:
            fill = PatternFill(start_color=color_map[year], end_color=color_map[year], fill_type="solid")
            for c in row:
                c.fill = fill

wb.save(filename)

# Визуелизација на температурата (Целзиусови степени)
plt.figure(figsize=(10, 5))
df_plot = df_history.sort_values("datum")
plt.plot(df_plot["datum"], df_plot["temperature_2m_max"], label="Максимална температура (°C)")
plt.plot(df_plot["datum"], df_plot["temperature_2m_min"], label="Минимална температура (°C)")
plt.xlabel("Датум")
plt.ylabel("Температура (°C)")
plt.title(f"Температурен тренд во {city_name} ({start_date} до {end_date})")
plt.legend()
plt.tight_layout()
plt.show()

print("Скриптата е успешно завршена.")
